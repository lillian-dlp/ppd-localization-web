#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
解析“多语言字符串总表”Excel, 产出可供 PPD 本地化比对使用的结构。

总表结构(每个工作表一份驱动字符串):
  序号 | 字符限制 | 简体中文 | 英语(en) | 俄语(ru) | 德语(de) | ... (每语言一列)
行按字符串文本编号(不含 PPD 选项键), 故用英文/中文列做文本锚点与 PPD 连接。

用法:
  作模块:  from xlsx_baseline import parse_all
  CLI:    python3 xlsx_baseline.py <xlsx> [sheet] [target_lang]
"""

import io
import re
import sys
import json

import openpyxl


# 表头括号内语言码的归一映射(处理脏码)
_CODE_FIX = {
    'enus': 'en', 'en': 'en',
    'esmx': 'es', 'es': 'es',
    'zhcn': 'zh_CN', 'zh_cn': 'zh_CN',
    'zhtw': 'zh_TW', 'zh_tw': 'zh_TW',
    'bu': 'bg', 'bg': 'bg',
    'kz': 'kk', 'kk': 'kk',
}
# 无括号码时按中文语言名归一
_NAME_MAP = [
    ('简体中文', 'zh_CN'), ('繁体中文', 'zh_TW'),
    ('英语', 'en'), ('英文', 'en'),
    ('韩国', 'ko'), ('韩语', 'ko'),
    ('哈萨克斯坦', 'kk'),
    ('中文', 'zh_CN'),  # 放最后, 避免抢先于“简体/繁体中文”
]


def header_to_code(header):
    """把一个语言列表头归一成语言码; 非语言列返回 None。"""
    if header is None:
        return None
    h = str(header).replace('\n', '').replace('\r', '').strip()
    if not h:
        return None
    if '序号' in h or '字符限制' in h or '限制' in h:
        return None
    # 括号内的码
    m = re.search(r'[（(]\s*([A-Za-z][A-Za-z_]*)\s*[)）]', h)
    if m:
        raw = m.group(1).lower()
        return _CODE_FIX.get(raw, raw)
    # 否则按中文名
    for name, code in _NAME_MAP:
        if name in h:
            return code
    return None


def _clean(v):
    if v is None:
        return ''
    return str(v).replace('\r', '').replace('\n', ' ').strip()


def parse_sheet(ws):
    """解析一个工作表 -> {name, languages, has_limit, rows}。"""
    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        return {'name': ws.title, 'languages': [], 'has_limit': False, 'rows': []}
    header = rows_iter[0]

    lang_cols = []   # [(col_idx, code)]
    en_col = zh_col = limit_col = None
    for i, h in enumerate(header):
        hn = '' if h is None else str(h).replace('\n', '').replace('\r', '').strip()
        if '字符限制' in hn or hn == '限制':
            limit_col = i
            continue
        code = header_to_code(h)
        if code is None:
            continue
        lang_cols.append((i, code))
        if code == 'en' and en_col is None:
            en_col = i
        if code == 'zh_CN' and zh_col is None:
            zh_col = i

    out_rows = []
    for r in rows_iter[1:]:
        def cell(idx):
            return _clean(r[idx]) if idx is not None and idx < len(r) else ''
        texts = {}
        for ci, code in lang_cols:
            t = cell(ci)
            if t:
                texts[code] = t
        if not texts:
            continue  # 整行空
        limit = None
        if limit_col is not None and limit_col < len(r) and r[limit_col] not in (None, ''):
            try:
                limit = int(float(str(r[limit_col]).strip()))
            except (ValueError, TypeError):
                limit = None
        out_rows.append({
            'en': cell(en_col),
            'zh': cell(zh_col),
            'limit': limit,
            'texts': texts,
        })

    return {
        'name': ws.title,
        'languages': [c for _, c in lang_cols],
        'has_limit': limit_col is not None,
        'rows': out_rows,
    }


def load_workbook(path_or_bytes):
    if isinstance(path_or_bytes, (bytes, bytearray)):
        src = io.BytesIO(path_or_bytes)
    else:
        src = path_or_bytes
    return openpyxl.load_workbook(src, read_only=True, data_only=True)


def parse_all(path_or_bytes):
    """解析整个工作簿 -> {sheets:[parse_sheet, ...]}。"""
    wb = load_workbook(path_or_bytes)
    sheets = []
    for ws in wb.worksheets:
        s = parse_sheet(ws)
        if s['languages']:          # 只保留含语言列的表
            sheets.append(s)
    return {'sheets': sheets}


def _cli():
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    path = sys.argv[1]
    want_sheet = sys.argv[2] if len(sys.argv) > 2 else None
    want_lang = sys.argv[3] if len(sys.argv) > 3 else None
    data = parse_all(path)
    for s in data['sheets']:
        if want_sheet and s['name'] != want_sheet:
            continue
        print('== 工作表 %s | 语言 %d 种: %s | 字符限制列: %s | 数据 %d 行 =='
              % (s['name'], len(s['languages']), ','.join(s['languages']),
                 '有' if s['has_limit'] else '无', len(s['rows'])))
        if want_lang:
            n = 0
            for row in s['rows']:
                t = row['texts'].get(want_lang)
                if t and row['en']:
                    print('  [%s] EN<%s>  ->  %s<%s>%s'
                          % ((str(row['limit']) if row['limit'] else '-'),
                             row['en'], want_lang, t,
                             '' if row['limit'] is None else '  (上限%d)' % row['limit']))
                    n += 1
            print('  目标语言 %s 可连接条目: %d' % (want_lang, n))
    return 0


if __name__ == '__main__':
    sys.exit(_cli())
